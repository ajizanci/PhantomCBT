import openpyxl
from django.contrib.auth.models import User
from user.models import Profile
from .models import Question, Option
from random import randint


def generate_unique_id():
    uid = randint(100000, 999999)
    try:
        existing_student = Profile.objects.get(unique_id=uid)
    except Profile.DoesNotExist:
        return uid

    return generate_unique_id()


def create_questions(workbook, exam):
    questions = None
    wb = openpyxl.load_workbook(workbook)
    sheet = wb.get_sheet_by_name("Questions")
    questions = get_questions_from_sheet(sheet)

    for question in questions:
        q = Question.objects.create(
            examination=exam, content=question["content"])
        for option in question["options"]:
            is_correct = option == question["correct_option"]
            Option.objects.create(question=q, option_content=option, is_correct=is_correct)

    return True


def create_students(workbook, exam):
    students = None
    try:
        wb = openpyxl.load_workbook(workbook)
        sheet = wb.get_sheet_by_name("Students")
        students = get_students_from_sheet(sheet)
    except:
        return False

    for student in students:
        uid = generate_unique_id()
        st = User.objects.create_user(username=f"{student['first_name']}{uid}", first_name=student["first_name"], last_name=student["last_name"])
        Profile.objects.create(user=st, account_type=2, unique_id=uid, examination=exam)

    return True


def get_students_from_sheet(sheet):
    students = []
    for row in range(2, sheet.max_row + 1):
        first_name = sheet["A"+str(row)].value
        last_name = sheet["B"+str(row)].value
        students.append({'first_name': first_name, 'last_name': last_name})

    return students


def get_questions_from_sheet(sheet):
    questions = []
    for row in range(2, sheet.max_row + 1):
        content = sheet["A" + str(row)].value
        options = map(lambda x: x.strip(), sheet["B" + str(row)].value.split(","))
        correct_option = str(sheet["C" + str(row)].value).strip()
        print(correct_option)
        questions.append({
            'content': content,
            'options': options,
            'correct_option': correct_option
        })

    return questions
